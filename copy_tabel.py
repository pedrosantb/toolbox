import openpyxl as xl
from openpyxl.workbook import Workbook, workbook

import argparse 

class copytable:
    def __init__(self, workbook):
        self.workbook = workbook

    def copysh(self, newWorkbook, sheet):

        self.newWorkbook = newWorkbook
        self.sheet = sheet

        wb1 = xl.load_workbook(self.workbook)
        ws1 = wb1.worksheets[self.sheet]

        wb2 = xl.Workbook()
        ws2 = wb2.create_sheet(ws1.title)

        for row in ws1:
            for cell in row:
                
                ws2[cell.coordinate].value = cell.value
        
        del wb2['Sheet']
        wb2.save(self.newWorkbook)
    
    def copywb(self, newWorkbook):
        self.newWorkbook = newWorkbook

        wb1 = xl.load_workbook((self.workbook))
        wb2 = xl.Workbook()

        for sheet in wb1.sheetnames: 
            ws1 = wb1[sheet]
            ws2 = wb2.create_sheet(ws1.title)

            for row in ws1:
                for cell in row:
                    ws2[cell.coordinate].value = cell.value
        
        del wb2['Sheet']
        wb2.save(self.newWorkbook)


def copysheet():

        tableName = str(input('Insert the original table filepath: '))
        sheetIndex = int(input('Insert the sheet Index (starting at 0): '))
        print("\n")
        newTable = str(input('Insert the new table filepath: '))

        try:
            name = copytable(str(tableName))
            name.copysh(str(newTable), sheetIndex)
            
            return "Sheet copied successfully!"
        
        except:

            return "An error has occurred, try to check the filepath"
     

def copytb():

    tableName = str(input('Insert the original table filepath: '))
    print("\n")
    newTable = str(input('Insert the new table filepath: '))

    try:
        name = copytable(str(tableName))
        name.copywb(str(newTable))

        return "Table copied successfully!"
    except:
        return "An error has occurred, try to check the filepath"
    

if __name__ == '__main__':
    
    print('''
          ____                    _____     _     _      
        /  ___|___  _ __  _   _  |_   _|_ _| |__ | | ___ 
        | |   / _ \| '_ \| | | |   | |/ _` | '_ \| |/ _ |
        | |__| (_) | |_) | |_| |   | | (_| | |_) | |  __/
        \____ \___/| .__/ \__, |   |_|\__,_|_.__/|_|\___|
                   |_|    |___/                          
        
        BY: Pedro StªBárbara
    ''')

    parser = argparse.ArgumentParser("copytable")
    parser.add_argument("Action", help="Use copytable or copysheet", type=str)
    args = parser.parse_args()

    if args.Action == 'copytable':
        print(copytb())

    elif args.Action == 'copysheet':
        print(copysheet())

    else:
        print('Invalid argument!')
        print('try copytable or copysheet')
            